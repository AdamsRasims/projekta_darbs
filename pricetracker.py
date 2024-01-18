import selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import time
from openpyxl import Workbook, load_workbook
import tldextract

wb = load_workbook('tel_linki.xlsx')
ws = wb.active
max_row = ws.max_row

service = Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=option)

def telefons():
    for row in range(6, max_row+1):
        url = ws['b' + str(row)].value
        XP1 = ws['e' + str(row)].value
        XP2 = ws['f' + str(row)].value

        nosaukums = tldextract.extract(url)
        tirgotajs = nosaukums.domain.capitalize()
        ws['a' + str(row)] = tirgotajs

        if(row>8):
            if(ws['e' + str(row)].value == None):
                irxpath=False
                for row2 in range (2,9):
                    url2 = ws['b' + str(row2)].value
                    nosaukums2 = tldextract.extract(url2)
                    if nosaukums.domain == nosaukums2.domain:
                        XP1 = ws['e' + str(row2)].value
                        irxpath=True
                        break



        if(ws['f' + str(row)].value == None):
            irxpath=False
            for row2 in range (2,9):
                url2 = ws['b' + str(row2)].value
                nosaukums2 = tldextract.extract(url2)
                if nosaukums.domain == nosaukums2.domain:
                    XP2 = ws['f' + str(row2)].value
                    irxpath=True
                    break
                if irxpath:
                    break


        driver.get(url)
        time.sleep(4)

        if XP1 != None:  
            driver.find_element(By.XPATH, XP1).click()
            time.sleep(3)
        
        cena = driver.find_element(By.XPATH, XP2)
        
        cena2 = cena.text

        if ("," in cena2) or ("." in cena2):
            pass
        else:
            cena2 = cena2 + ",00"

        cena3 = cena2.replace(" ", "").replace(".", ",").replace("€", "") + ' EUR'
        ws['c' + str(row)] = cena3
    
    wb.save('tel_linki.xlsx')

telefons()
wb.close()
